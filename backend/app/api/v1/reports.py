import csv
import io
from datetime import UTC, date, datetime, time, timedelta
from typing import Annotated

import openpyxl
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import aliased
from xhtml2pdf import pisa

from app.core.dependencies import AdminOnly, CurrentUser, NotStudent, get_db
from app.models.attendance import AttendanceRecord, Session
from app.models.eligibility import AthleteEligibility, EligibilityStatus
from app.models.facility import Facility
from app.models.incident import Incident
from app.models.inventory import (
    BorrowTransaction,
    BorrowTransactionItem,
    Equipment,
    EquipmentCondition,
    TransactionStatus,
)
from app.models.reservation import VenueReservationRequest
from app.models.sanction import Sanction, SanctionStatus
from app.models.user import User
from app.schemas.attendance import AttendanceReportFilter
from app.services.audit_service import audit_log
from app.services.report_service import ReportService

router = APIRouter()


async def _log_report_export(
    db: AsyncSession,
    user,
    report: str,
    period: str,
    fmt: str,
    target: str,
) -> None:
    if fmt == "json":
        return
    await audit_log(
        db=db,
        action="REPORT_GENERATED",
        module="Reports",
        description=f"Exported {report} report ({period}, {fmt.upper()})",
        resource_type="Report",
        details={"format": fmt, "report": report, "period": period, "target": target},
        current_user=user,
    )
    await db.commit()


@router.get(
    "/attendance/pdf",
    summary="Generate attendance report PDF",
    response_class=StreamingResponse,
)
async def attendance_pdf(
    _user: NotStudent,
    db: Annotated[AsyncSession, Depends(get_db)],
    sport_or_art: str | None = Query(None),
    session_id: str | None = Query(None),
    date_from: datetime | None = Query(None),
    date_to: datetime | None = Query(None),
) -> StreamingResponse:
    report_service = ReportService(db)
    pdf_bytes = await report_service.generate_attendance_pdf(
        sport_or_art=sport_or_art,
        date_from=date_from,
        date_to=date_to,
    )
    await audit_log(
        db=db,
        action="REPORT_GENERATED",
        module="Reports",
        description="Generated attendance report (PDF)",
        resource_type="Report",
        details={"format": "pdf", "report": "attendance", "sport_or_art": sport_or_art,
                 "date_from": str(date_from) if date_from else None, "date_to": str(date_to) if date_to else None},
        current_user=_user,
    )
    await db.commit()
    return StreamingResponse(
        iter([pdf_bytes]),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=attendance_report.pdf"},
    )


@router.get(
    "/attendance/xlsx",
    summary="Export attendance report XLSX",
    response_class=StreamingResponse,
)
async def attendance_xlsx(
    _user: NotStudent,
    db: Annotated[AsyncSession, Depends(get_db)],
    sport_or_art: str | None = Query(None),
    date_from: datetime | None = Query(None),
    date_to: datetime | None = Query(None),
) -> StreamingResponse:
    report_service = ReportService(db)
    xlsx_bytes = await report_service.generate_attendance_xlsx(
        sport_or_art=sport_or_art,
        date_from=date_from,
        date_to=date_to,
    )
    await audit_log(
        db=db,
        action="REPORT_GENERATED",
        module="Reports",
        description="Exported attendance report (XLSX)",
        resource_type="Report",
        details={"format": "xlsx", "report": "attendance", "sport_or_art": sport_or_art,
                 "date_from": str(date_from) if date_from else None, "date_to": str(date_to) if date_to else None},
        current_user=_user,
    )
    await db.commit()
    return StreamingResponse(
        iter([xlsx_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=attendance_report.xlsx"},
    )


@router.get(
    "/inventory/pdf",
    summary="Generate inventory report PDF",
    response_class=StreamingResponse,
)
async def inventory_pdf(
    _user: NotStudent,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> StreamingResponse:
    report_service = ReportService(db)
    pdf_bytes = await report_service.generate_inventory_pdf()
    await audit_log(
        db=db,
        action="REPORT_GENERATED",
        module="Reports",
        description="Generated inventory report (PDF)",
        resource_type="Report",
        details={"format": "pdf", "report": "inventory"},
        current_user=_user,
    )
    await db.commit()
    return StreamingResponse(
        iter([pdf_bytes]),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=inventory_report.pdf"},
    )


@router.get(
    "/inventory/xlsx",
    summary="Export inventory report XLSX",
    response_class=StreamingResponse,
)
async def inventory_xlsx(
    _user: NotStudent,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> StreamingResponse:
    report_service = ReportService(db)
    xlsx_bytes = await report_service.generate_inventory_xlsx()
    await audit_log(
        db=db,
        action="REPORT_GENERATED",
        module="Reports",
        description="Exported inventory report (XLSX)",
        resource_type="Report",
        details={"format": "xlsx", "report": "inventory"},
        current_user=_user,
    )
    await db.commit()
    return StreamingResponse(
        iter([xlsx_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory_report.xlsx"},
    )


@router.get(
    "/dashboard/summary",
    summary="Dashboard summary data",
)
async def dashboard_summary(
    _user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict:
    report_service = ReportService(db)
    return await report_service.get_dashboard_summary()


@router.get(
    "/inventory/monthly",
    summary="Monthly inventory summary (Admin / Staff)",
    response_model=None,
)
async def inventory_monthly(
    _user: AdminOnly,
    db: Annotated[AsyncSession, Depends(get_db)],
    year: int = Query(..., ge=2020, le=2100, description="Report year"),
    month: int = Query(..., ge=1, le=12, description="Report month (1-12)"),
    format: str = Query("json", pattern="^(json|pdf|xlsx|csv)$", description="Response format"),
) -> StreamingResponse | dict:
    report_service = ReportService(db)
    result = await report_service.generate_inventory_monthly_report(year=year, month=month)

    if format == "csv":
        csv_data = _build_monthly_csv(result)
        return StreamingResponse(
            iter([csv_data]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=inventory_monthly_{year}_{month:02d}.csv"},
        )
    if format == "pdf":
        pdf_bytes = await report_service.render_monthly_report_pdf(result)
        filename = f"inventory_monthly_{year}_{month:02d}.pdf"
        return StreamingResponse(
            iter([pdf_bytes]),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    if format == "xlsx":
        xlsx_bytes = await report_service.render_monthly_report_xlsx(result)
        filename = f"inventory_monthly_{year}_{month:02d}.xlsx"
        return StreamingResponse(
            iter([xlsx_bytes]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    return result


async def _fetch_attendance_logs(
    db: AsyncSession,
    date_from: datetime,
    date_to: datetime,
    sport_or_art: str | None = None,
) -> list[dict]:
    query = (
        select(AttendanceRecord, User, Session)
        .join(User, AttendanceRecord.student_id == User.id)
        .join(Session, AttendanceRecord.session_id == Session.id)
        .where(AttendanceRecord.time_in >= date_from, AttendanceRecord.time_in <= date_to)
    )
    if sport_or_art:
        query = query.where(Session.sport_or_art == sport_or_art)
    query = query.order_by(AttendanceRecord.time_in.desc())

    result = await db.execute(query)
    rows = []
    for record, user, session in result.all():
        rows.append({
            "id": str(record.id),
            "student_name": user.full_name,
            "student_id": user.student_id or "",
            "student_role": user.role.value if hasattr(user.role, 'value') else str(user.role),
            "student_email": user.email,
            "session_name": session.name,
            "sport_or_art": session.sport_or_art or "",
            "activity_type": session.activity_type.value,
            "time_in": record.time_in.isoformat() if record.time_in else "",
            "time_out": record.time_out.isoformat() if record.time_out else "",
            "attendance_date": record.time_in.strftime("%Y-%m-%d") if record.time_in else "",
            "duration_minutes": record.duration_minutes,
            "status": record.status or "",
            "confidence": record.time_in_confidence,
            "is_complete": record.is_complete,
        })
    return rows


def _build_attendance_csv(rows: list[dict]) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Full Name", "Student ID", "Role", "Sport / Art",
        "Attendance Status", "Time In", "Time Out", "Attendance Date",
    ])
    for r in rows:
        writer.writerow([
            r["student_name"], r["student_id"], r["student_role"],
            r["sport_or_art"], r["status"], r["time_in"], r["time_out"],
            r["attendance_date"],
        ])
    output.seek(0)
    return output.getvalue()


def _build_attendance_xlsx(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Logs"

    headers = [
        "Full Name", "Student ID", "Role", "Sport / Art",
        "Attendance Status", "Time In", "Time Out", "Attendance Date",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="1E3A5F")
        cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)

    for i, r in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=r["student_name"])
        ws.cell(row=i, column=2, value=r["student_id"])
        ws.cell(row=i, column=3, value=r["student_role"])
        ws.cell(row=i, column=4, value=r["sport_or_art"])
        ws.cell(row=i, column=5, value=r["status"])
        ws.cell(row=i, column=6, value=r["time_in"])
        ws.cell(row=i, column=7, value=r["time_out"])
        ws.cell(row=i, column=8, value=r["attendance_date"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _build_attendance_html(rows: list[dict], title: str, period: str) -> str:
    rows_html = ""
    for r in rows:
        rows_html += f"""
        <tr>
            <td>{r['student_name']}</td>
            <td>{r['student_id']}</td>
            <td>{r['student_role']}</td>
            <td>{r['sport_or_art']}</td>
            <td>{r['status']}</td>
            <td>{_fmt_iso_time(r['time_in'])}</td>
            <td>{_fmt_iso_time(r['time_out'])}</td>
            <td>{_fmt_iso_date(r['attendance_date'])}</td>
        </tr>"""

    return f"""
    <html><head><meta charset="UTF-8">
    <style>
        @page {{ size: A4 landscape; margin: 12mm 9mm; }}
        body {{ font-family: Helvetica, sans-serif; font-size: 9pt; }}
        h1 {{ color: #1E3A5F; font-size: 14pt; }}
        p {{ color: #666; font-size: 8pt; }}
        table {{ width: 100%; table-layout: fixed; border-collapse: collapse; margin-top: 10px; }}
        th {{ background: #1E3A5F; color: white; padding: 5px 6px; text-align: left; font-size: 8pt; word-wrap: break-word; }}
        td {{ padding: 4px 6px; border-bottom: 1px solid #ddd; font-size: 8pt; word-wrap: break-word; }}
    </style></head><body>
        <h1>OSCA — {title}</h1>
        <p>{period} | Records: {len(rows)} | Generated: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}</p>
        <table>
            <tr><th>Name</th><th>ID</th><th>Role</th><th>Sport/Art</th><th>Status</th><th>Time In</th><th>Time Out</th><th>Date</th></tr>
            {rows_html}
        </table>
    </body></html>"""


@router.get(
    "/attendance/daily",
    response_model=None,
    summary="Daily attendance log",
)
async def daily_attendance(
    _user: NotStudent,
    db: Annotated[AsyncSession, Depends(get_db)],
    log_date: date | None = Query(None, description="Date (default: today)"),
    sport_or_art: str | None = Query(None),
    format: str = Query("json", pattern="^(json|csv|xlsx|pdf)$"),
) -> StreamingResponse | list:
    target = log_date or date.today()
    date_from = datetime(target.year, target.month, target.day, tzinfo=UTC)
    date_to = date_from + timedelta(days=1) - timedelta(seconds=1)

    rows = await _fetch_attendance_logs(db, date_from, date_to, sport_or_art)
    period_str = f"Daily Attendance — {target.isoformat()}"
    await _log_report_export(db, _user, "attendance", "Daily", format, target.isoformat())

    if format == "csv":
        csv_data = _build_attendance_csv(rows)
        return StreamingResponse(
            iter([csv_data]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=daily-attendance-{target.isoformat()}.csv"},
        )
    if format == "xlsx":
        xlsx_data = _build_attendance_xlsx(rows)
        return StreamingResponse(
            iter([xlsx_data]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=daily-attendance-{target.isoformat()}.xlsx"},
        )
    if format == "pdf":
        html = _build_attendance_html(rows, period_str, period_str)
        pdf_bytes = io.BytesIO()
        pisa.CreatePDF(io.StringIO(html), dest=pdf_bytes)
        pdf_bytes.seek(0)
        return StreamingResponse(
            pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=daily-attendance-{target.isoformat()}.pdf"},
        )

    return rows


@router.get(
    "/attendance/weekly",
    response_model=None,
    summary="Weekly attendance log",
)
async def weekly_attendance(
    _user: NotStudent,
    db: Annotated[AsyncSession, Depends(get_db)],
    week_start: date | None = Query(None, description="Start of week (Monday, default: current week)"),
    sport_or_art: str | None = Query(None),
    format: str = Query("json", pattern="^(json|csv|xlsx|pdf)$"),
) -> StreamingResponse | list:
    start = week_start or (date.today() - timedelta(days=date.today().weekday()))
    date_from = datetime(start.year, start.month, start.day, tzinfo=UTC)
    date_to = date_from + timedelta(days=7) - timedelta(seconds=1)

    rows = await _fetch_attendance_logs(db, date_from, date_to, sport_or_art)
    period_str = f"Weekly Attendance — {start.isoformat()} to {(start + timedelta(days=6)).isoformat()}"
    await _log_report_export(db, _user, "attendance", "Weekly", format, f"{start.isoformat()}_to_{start.isoformat()}")

    if format == "csv":
        csv_data = _build_attendance_csv(rows)
        return StreamingResponse(
            iter([csv_data]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=weekly-attendance-{start.isoformat()}.csv"},
        )
    if format == "xlsx":
        xlsx_data = _build_attendance_xlsx(rows)
        return StreamingResponse(
            iter([xlsx_data]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=weekly-attendance-{start.isoformat()}.xlsx"},
        )
    if format == "pdf":
        html = _build_attendance_html(rows, period_str, period_str)
        pdf_bytes = io.BytesIO()
        pisa.CreatePDF(io.StringIO(html), dest=pdf_bytes)
        pdf_bytes.seek(0)
        return StreamingResponse(
            pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=weekly-attendance-{start.isoformat()}.pdf"},
        )

    return rows


@router.get(
    "/attendance/monthly",
    response_model=None,
    summary="Monthly attendance log",
)
async def monthly_attendance(
    _user: NotStudent,
    db: Annotated[AsyncSession, Depends(get_db)],
    year: int = Query(None, ge=2020, le=2100),
    month: int = Query(None, ge=1, le=12),
    sport_or_art: str | None = Query(None),
    format: str = Query("json", pattern="^(json|csv|xlsx|pdf)$"),
) -> StreamingResponse | list:
    today = date.today()
    y = year or today.year
    m = month or today.month
    from calendar import monthrange
    _, last_day = monthrange(y, m)
    date_from = datetime(y, m, 1, tzinfo=UTC)
    date_to = datetime(y, m, last_day, 23, 59, 59, tzinfo=UTC)

    rows = await _fetch_attendance_logs(db, date_from, date_to, sport_or_art)
    from calendar import month_name
    period_str = f"Monthly Attendance — {month_name[m]} {y}"
    await _log_report_export(db, _user, "attendance", "Monthly", format, f"{y}-{m:02d}")

    if format == "csv":
        csv_data = _build_attendance_csv(rows)
        return StreamingResponse(
            iter([csv_data]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=monthly-attendance-{y}-{m:02d}.csv"},
        )
    if format == "xlsx":
        xlsx_data = _build_attendance_xlsx(rows)
        return StreamingResponse(
            iter([xlsx_data]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=monthly-attendance-{y}-{m:02d}.xlsx"},
        )
    if format == "pdf":
        html = _build_attendance_html(rows, period_str, period_str)
        pdf_bytes = io.BytesIO()
        pisa.CreatePDF(io.StringIO(html), dest=pdf_bytes)
        pdf_bytes.seek(0)
        return StreamingResponse(
            pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=monthly-attendance-{y}-{m:02d}.pdf"},
        )

    return rows


def _fmt_val(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        if v.tzinfo is not None:
            v = v.astimezone(UTC)
        return v.strftime("%B %d, %Y %I:%M %p")
    if isinstance(v, date):
        return v.strftime("%B %d, %Y")
    if isinstance(v, time):
        return v.strftime("%I:%M %p")
    if isinstance(v, bool):
        return "Yes" if v else "No"
    if hasattr(v, "value"):
        return v.value
    return str(v)


def _fmt_iso_time(value: str) -> str:
    if not value:
        return "—"
    try:
        dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return value
    return dt.strftime("%I:%M %p")


def _fmt_iso_date(value: str) -> str:
    if not value:
        return "—"
    try:
        d = date.fromisoformat(value[:10])
    except ValueError:
        return value
    return d.strftime("%B %d, %Y")


def _xlsx_val(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.astimezone(UTC).replace(tzinfo=None) if v.tzinfo is not None else v
    if isinstance(v, (date, time)):
        return v
    if hasattr(v, "value"):
        return v.value
    return v


def _rows_to_csv(rows: list[dict], headers: list[str]) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for r in rows:
        writer.writerow([_fmt_val(r.get(h)) for h in headers])
    output.seek(0)
    return output.getvalue()


def _rows_to_xlsx(rows: list[dict], headers: list[str], title: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    sheet_name = (title or "Report")[:31] or "Report"
    for ch in "\\/*?:[]":
        sheet_name = sheet_name.replace(ch, " ")
    ws.title = sheet_name
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="1E3A5F")
        cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
    for i, r in enumerate(rows, 2):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=i, column=col, value=_xlsx_val(r.get(h)))
            if i % 2 == 0:
                cell.fill = openpyxl.styles.PatternFill("solid", fgColor="EBF0F7")
    for column in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in column), default=10)
        ws.column_dimensions[column[0].column_letter].width = min(max_len + 4, 50)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _rows_to_pdf(rows: list[dict], headers: list[str], title: str, subtitle: str) -> bytes:
    thead = "".join(f"<th>{h}</th>" for h in headers)
    body = ""
    for i, r in enumerate(rows):
        bg = "#EBF0F7" if i % 2 == 0 else "#FFFFFF"
        cells = "".join(f"<td>{_fmt_val(r.get(h)) or '—'}</td>" for h in headers)
        body += f'<tr style="background:{bg}">{cells}</tr>'
    html = f"""
    <!DOCTYPE html><html><head><meta charset="UTF-8">
    <style>
        @page {{ size: A4 landscape; margin: 12mm 9mm; }}
        body {{ font-family: Arial, sans-serif; font-size: 9pt; }}
        h1 {{ color: #1E3A5F; font-size: 14pt; }}
        p {{ color: #666; font-size: 8pt; }}
        table {{ width: 100%; table-layout: fixed; border-collapse: collapse; margin-top: 10px; }}
        th {{ background: #1E3A5F; color: white; padding: 5px 6px; text-align: left; font-size: 8pt; word-wrap: break-word; }}
        td {{ padding: 4px 6px; border-bottom: 1px solid #DDD; font-size: 8pt; word-wrap: break-word; }}
    </style></head><body>
    <h1>NAAP-Villamor OSCA — {title}</h1>
    <p>{subtitle} | Records: {len(rows)} | Generated: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}</p>
    <table><tr>{thead}</tr>{body}</table>
    </body></html>"""
    buffer = io.BytesIO()
    pisa.CreatePDF(io.StringIO(html), dest=buffer)
    return buffer.getvalue()


async def _export_rows(
    db: AsyncSession,
    user,
    rows: list[dict],
    headers: list[str],
    title: str,
    period: str,
    fmt: str,
    report_name: str,
    slug: str,
    date_from=None,
    date_to=None,
) -> StreamingResponse | list:
    if fmt == "json":
        return rows
    await _log_report_export(db, user, report_name, period, fmt, f"{date_from or 'all'} – {date_to or 'all'}")
    if fmt == "csv":
        return StreamingResponse(
            iter([_rows_to_csv(rows, headers)]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={slug}.csv"},
        )
    if fmt == "xlsx":
        data = _rows_to_xlsx(rows, headers, title)
        return StreamingResponse(
            iter([data]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={slug}.xlsx"},
        )
    data = _rows_to_pdf(rows, headers, title, period)
    return StreamingResponse(
        iter([data]),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={slug}.pdf"},
    )


def _day_start(d: date | None) -> datetime | None:
    return datetime(d.year, d.month, d.day, tzinfo=UTC) if d else None


def _day_end(d: date | None) -> datetime | None:
    return datetime(d.year, d.month, d.day, 23, 59, 59, tzinfo=UTC) if d else None


def _range_label(date_from, date_to) -> str:
    if date_from and date_to:
        return f"{date_from.isoformat()} to {date_to.isoformat()}"
    if date_from:
        return f"From {date_from.isoformat()}"
    if date_to:
        return f"Until {date_to.isoformat()}"
    return "All time"


def _build_monthly_csv(report: dict) -> str:
    from calendar import month_name
    period = report["period"]
    label = f"{month_name[period['month']]} {period['year']}"
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Period", "Metric", "Value"])
    for group, metric, value in [
        ("Summary", "Active Equipment", report["total_active_equipment"]),
        ("Summary", "Borrowed This Month", report["borrowed_this_month"]),
        ("Summary", "Returned This Month", report["returned_this_month"]),
        ("Summary", "Overdue at End of Month", report["overdue_at_end_of_month"]),
    ]:
        writer.writerow([label, f"{group} — {metric}", value])
    for i, e in enumerate(report["top_5_borrowed"], 1):
        writer.writerow([label, f"Top 5 Borrowed #{i}", f"{e['name']} — {e['borrow_count']}"])
    for cond, count in report["condition_breakdown"].items():
        writer.writerow([label, f"Condition — {cond.title()}", count])
    output.seek(0)
    return output.getvalue()


async def _fetch_inventory_items_map(db: AsyncSession, transaction_ids: list) -> dict:
    if not transaction_ids:
        return {}
    result = await db.execute(
        select(BorrowTransactionItem.transaction_id, Equipment.name, BorrowTransactionItem.quantity)
        .join(Equipment, Equipment.id == BorrowTransactionItem.equipment_id)
        .where(BorrowTransactionItem.transaction_id.in_(transaction_ids))
    )
    mapping: dict = {}
    for txn_id, name, qty in result.all():
        mapping.setdefault(txn_id, []).append(f"{name} (x{qty})")
    return mapping


async def _fetch_borrowing_history(db: AsyncSession, date_from, date_to) -> list[dict]:
    query = (
        select(BorrowTransaction, User)
        .join(User, BorrowTransaction.instructor_id == User.id)
    )
    if date_from:
        query = query.where(BorrowTransaction.borrowed_at >= date_from)
    if date_to:
        query = query.where(BorrowTransaction.borrowed_at <= date_to)
    query = query.order_by(BorrowTransaction.borrowed_at.desc())
    result = await db.execute(query)
    pairs = result.all()
    items_map = await _fetch_inventory_items_map(db, [t.id for t, _ in pairs])
    rows = []
    for t, instructor in pairs:
        rows.append({
            "Transaction": f"TXN-{t.id.hex[:12].upper()}",
            "Instructor": instructor.full_name,
            "Items": ", ".join(items_map.get(t.id, [])),
            "Borrowed At": t.borrowed_at,
            "Expected Return": t.expected_return,
            "Returned At": t.returned_at,
            "Status": t.status.value,
            "Notes": t.notes,
        })
    return rows


async def _fetch_returned_equipment(db: AsyncSession, date_from, date_to) -> list[dict]:
    query = (
        select(BorrowTransaction, User)
        .join(User, BorrowTransaction.instructor_id == User.id)
        .where(BorrowTransaction.status == TransactionStatus.RETURNED)
    )
    if date_from:
        query = query.where(BorrowTransaction.returned_at >= date_from)
    if date_to:
        query = query.where(BorrowTransaction.returned_at <= date_to)
    query = query.order_by(BorrowTransaction.returned_at.desc())
    result = await db.execute(query)
    pairs = result.all()
    items_map = await _fetch_inventory_items_map(db, [t.id for t, _ in pairs])
    rows = []
    for t, instructor in pairs:
        days = (t.returned_at - t.borrowed_at).days if t.returned_at else ""
        rows.append({
            "Transaction": f"TXN-{t.id.hex[:12].upper()}",
            "Instructor": instructor.full_name,
            "Items": ", ".join(items_map.get(t.id, [])),
            "Returned At": t.returned_at,
            "Days Borrowed": days,
            "Notes": t.notes,
        })
    return rows


async def _fetch_lost_damaged(db: AsyncSession, date_from, date_to) -> list[dict]:
    query = (
        select(BorrowTransactionItem, Equipment, BorrowTransaction, User)
        .join(Equipment, Equipment.id == BorrowTransactionItem.equipment_id)
        .join(BorrowTransaction, BorrowTransaction.id == BorrowTransactionItem.transaction_id)
        .join(User, BorrowTransaction.instructor_id == User.id)
        .where(or_(
            BorrowTransactionItem.return_condition == EquipmentCondition.POOR,
            BorrowTransactionItem.return_condition == EquipmentCondition.FOR_REPAIR,
            BorrowTransactionItem.return_condition == EquipmentCondition.CONDEMNED,
            BorrowTransactionItem.notes.ilike("%lost%"),
            BorrowTransactionItem.notes.ilike("%missing%"),
            BorrowTransactionItem.notes.ilike("%damaged%"),
        ))
    )
    if date_from:
        query = query.where(BorrowTransaction.returned_at >= date_from)
    if date_to:
        query = query.where(BorrowTransaction.returned_at <= date_to)
    query = query.order_by(BorrowTransaction.returned_at.desc())
    result = await db.execute(query)
    rows = []
    for item, equipment, txn, instructor in result.all():
        rows.append({
            "Transaction": f"TXN-{txn.id.hex[:12].upper()}",
            "Equipment": equipment.name,
            "Quantity": item.quantity,
            "Return Condition": item.return_condition.value if item.return_condition else "",
            "Instructor": instructor.full_name,
            "Returned At": txn.returned_at,
            "Notes": item.notes or txn.notes or "",
        })
    return rows


@router.get(
    "/inventory/equipment",
    response_model=None,
    summary="Equipment inventory report (all formats)",
)
async def equipment_inventory_report(
    _user: AdminOnly,
    db: Annotated[AsyncSession, Depends(get_db)],
    format: str = Query("json", pattern="^(json|csv|xlsx|pdf)$"),
) -> StreamingResponse | list:
    report_service = ReportService(db)
    rows = await report_service._fetch_all_equipment()
    headers = ["Equipment Name", "Category", "Condition", "QR Code",
               "Total Qty", "Available", "Borrowed", "Location", "Sport/Art"]
    return await _export_rows(db, _user, rows, headers, "Equipment Inventory Report",
                              "All time", format, "inventory-equipment", "inventory-equipment")


@router.get(
    "/inventory/borrowing-history",
    response_model=None,
    summary="Borrowing history report",
)
async def borrowing_history_report(
    _user: AdminOnly,
    db: Annotated[AsyncSession, Depends(get_db)],
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    format: str = Query("json", pattern="^(json|csv|xlsx|pdf)$"),
) -> StreamingResponse | list:
    rows = await _fetch_borrowing_history(db, _day_start(date_from), _day_end(date_to))
    headers = ["Transaction", "Instructor", "Items", "Borrowed At",
               "Expected Return", "Returned At", "Status", "Notes"]
    return await _export_rows(db, _user, rows, headers, "Borrowing History Report",
                              _range_label(date_from, date_to), format,
                              "borrowing-history", "borrowing-history", date_from, date_to)


@router.get(
    "/inventory/returned",
    response_model=None,
    summary="Returned equipment report",
)
async def returned_equipment_report(
    _user: AdminOnly,
    db: Annotated[AsyncSession, Depends(get_db)],
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    format: str = Query("json", pattern="^(json|csv|xlsx|pdf)$"),
) -> StreamingResponse | list:
    rows = await _fetch_returned_equipment(db, _day_start(date_from), _day_end(date_to))
    headers = ["Transaction", "Instructor", "Items", "Returned At", "Days Borrowed", "Notes"]
    return await _export_rows(db, _user, rows, headers, "Returned Equipment Report",
                              _range_label(date_from, date_to), format,
                              "returned-equipment", "returned-equipment", date_from, date_to)


@router.get(
    "/inventory/lost-damaged",
    response_model=None,
    summary="Lost / damaged equipment report",
)
async def lost_damaged_equipment_report(
    _user: AdminOnly,
    db: Annotated[AsyncSession, Depends(get_db)],
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    format: str = Query("json", pattern="^(json|csv|xlsx|pdf)$"),
) -> StreamingResponse | list:
    rows = await _fetch_lost_damaged(db, _day_start(date_from), _day_end(date_to))
    headers = ["Transaction", "Equipment", "Quantity", "Return Condition",
               "Instructor", "Returned At", "Notes"]
    return await _export_rows(db, _user, rows, headers, "Lost / Damaged Equipment Report",
                              _range_label(date_from, date_to), format,
                              "lost-damaged", "lost-damaged", date_from, date_to)


async def _fetch_venue_reservations(db: AsyncSession, date_from, date_to) -> list[dict]:
    query = (
        select(VenueReservationRequest, Facility, User)
        .join(Facility, Facility.id == VenueReservationRequest.facility_id)
        .join(User, User.id == VenueReservationRequest.requester_id)
    )
    if date_from:
        query = query.where(VenueReservationRequest.reservation_date >= date_from)
    if date_to:
        query = query.where(VenueReservationRequest.reservation_date <= date_to)
    query = query.order_by(VenueReservationRequest.reservation_date.desc(), VenueReservationRequest.start_time)
    result = await db.execute(query)
    rows = []
    for req, facility, requester in result.all():
        rows.append({
            "Venue": facility.name,
            "Requester": requester.full_name,
            "Purpose": req.purpose,
            "Date": req.reservation_date,
            "Start": req.start_time,
            "End": req.end_time,
            "Status": req.status.value,
            "Remarks": req.remarks,
            "Rejection Reason": req.rejection_reason,
        })
    return rows


async def _fetch_venue_usage(db: AsyncSession, date_from, date_to) -> list[dict]:
    query = (
        select(VenueReservationRequest, Facility)
        .join(Facility, Facility.id == VenueReservationRequest.facility_id)
    )
    if date_from:
        query = query.where(VenueReservationRequest.reservation_date >= date_from)
    if date_to:
        query = query.where(VenueReservationRequest.reservation_date <= date_to)
    result = await db.execute(query)
    stats: dict[str, dict] = {}
    for req, facility in result.all():
        entry = stats.setdefault(facility.name, {
            "Venue": facility.name,
            "Total Requests": 0,
            "Approved": 0,
            "Pending": 0,
            "Rejected": 0,
            "Approved Hours": 0.0,
        })
        entry["Total Requests"] += 1
        if req.status.value == "approved":
            entry["Approved"] += 1
            minutes = (req.end_time.hour * 60 + req.end_time.minute) - (req.start_time.hour * 60 + req.start_time.minute)
            entry["Approved Hours"] += round(minutes / 60, 1)
        elif req.status.value == "pending":
            entry["Pending"] += 1
        elif req.status.value == "rejected":
            entry["Rejected"] += 1
    return sorted(stats.values(), key=lambda x: x["Total Requests"], reverse=True)


async def _fetch_facility_status(db: AsyncSession) -> list[dict]:
    result = await db.execute(select(Facility).order_by(Facility.name))
    rows = []
    for f in result.scalars().all():
        rows.append({
            "Venue": f.name,
            "Status": f.status.value,
            "Capacity": f.capacity,
            "Active": f.is_active,
            "Description": f.description,
            "Last Updated": f.updated_at,
        })
    return rows


@router.get(
    "/facilities/venue-reservations",
    response_model=None,
    summary="Venue reservations report",
)
async def venue_reservations_report(
    _user: NotStudent,
    db: Annotated[AsyncSession, Depends(get_db)],
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    format: str = Query("json", pattern="^(json|csv|xlsx|pdf)$"),
) -> StreamingResponse | list:
    rows = await _fetch_venue_reservations(db, date_from, date_to)
    headers = ["Venue", "Requester", "Purpose", "Date", "Start", "End",
               "Status", "Remarks", "Rejection Reason"]
    return await _export_rows(db, _user, rows, headers, "Venue Reservations Report",
                              _range_label(date_from, date_to), format,
                              "venue-reservations", "venue-reservations", date_from, date_to)


@router.get(
    "/facilities/venue-usage",
    response_model=None,
    summary="Venue usage summary report",
)
async def venue_usage_report(
    _user: NotStudent,
    db: Annotated[AsyncSession, Depends(get_db)],
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    format: str = Query("json", pattern="^(json|csv|xlsx|pdf)$"),
) -> StreamingResponse | list:
    rows = await _fetch_venue_usage(db, date_from, date_to)
    headers = ["Venue", "Total Requests", "Approved", "Pending", "Rejected", "Approved Hours"]
    return await _export_rows(db, _user, rows, headers, "Venue Usage Report",
                              _range_label(date_from, date_to), format,
                              "venue-usage", "venue-usage", date_from, date_to)


@router.get(
    "/facilities/status",
    response_model=None,
    summary="Facility status report",
)
async def facility_status_report(
    _user: NotStudent,
    db: Annotated[AsyncSession, Depends(get_db)],
    format: str = Query("json", pattern="^(json|csv|xlsx|pdf)$"),
) -> StreamingResponse | list:
    rows = await _fetch_facility_status(db)
    headers = ["Venue", "Status", "Capacity", "Active", "Description", "Last Updated"]
    return await _export_rows(db, _user, rows, headers, "Facility Status Report",
                              "All time", format, "facility-status", "facility-status")


async def _fetch_eligibility(db: AsyncSession, statuses: list, date_from, date_to) -> list[dict]:
    query = (
        select(AthleteEligibility, User)
        .join(User, User.id == AthleteEligibility.student_id)
        .where(AthleteEligibility.status.in_(statuses))
    )
    if date_from:
        query = query.where(AthleteEligibility.start_date >= date_from)
    if date_to:
        query = query.where(AthleteEligibility.start_date <= date_to)
    query = query.order_by(User.last_name, User.first_name)
    result = await db.execute(query)
    rows = []
    for rec, student in result.all():
        rows.append({
            "Student Name": student.full_name,
            "Student ID": student.student_id or "",
            "Sport/Art": student.sport_or_art or "",
            "Status": rec.status.value,
            "Start Date": rec.start_date,
            "End Date": rec.end_date,
            "Medical Clearance": rec.medical_clearance,
            "Reason": rec.reason_type.value if rec.reason_type else "",
            "Reason Detail": rec.reason_detail,
            "Notes": rec.notes,
        })
    return rows


async def _eligibility_report(
    _user: NotStudent,
    db: AsyncSession,
    statuses: list,
    title: str,
    slug: str,
    date_from: date | None,
    date_to: date | None,
    format: str,
) -> StreamingResponse | list:
    rows = await _fetch_eligibility(db, statuses, date_from, date_to)
    headers = ["Student Name", "Student ID", "Sport/Art", "Status", "Start Date", "End Date",
               "Medical Clearance", "Reason", "Reason Detail", "Notes"]
    return await _export_rows(db, _user, rows, headers, title,
                              _range_label(date_from, date_to), format, slug, slug, date_from, date_to)


@router.get("/eligibility/eligible", response_model=None, summary="Eligible students report")
async def eligible_students_report(
    _user: NotStudent,
    db: Annotated[AsyncSession, Depends(get_db)],
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    format: str = Query("json", pattern="^(json|csv|xlsx|pdf)$"),
) -> StreamingResponse | list:
    return await _eligibility_report(_user, db, [EligibilityStatus.ELIGIBLE],
                                     "Eligible Students Report", "eligible-students",
                                     date_from, date_to, format)


@router.get("/eligibility/restricted", response_model=None, summary="Restricted students report")
async def restricted_students_report(
    _user: NotStudent,
    db: Annotated[AsyncSession, Depends(get_db)],
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    format: str = Query("json", pattern="^(json|csv|xlsx|pdf)$"),
) -> StreamingResponse | list:
    return await _eligibility_report(_user, db, [EligibilityStatus.RESTRICTED],
                                     "Restricted Students Report", "restricted-students",
                                     date_from, date_to, format)


@router.get("/eligibility/ineligible", response_model=None, summary="Ineligible students report")
async def ineligible_students_report(
    _user: NotStudent,
    db: Annotated[AsyncSession, Depends(get_db)],
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    format: str = Query("json", pattern="^(json|csv|xlsx|pdf)$"),
) -> StreamingResponse | list:
    return await _eligibility_report(_user, db, [EligibilityStatus.INELIGIBLE],
                                     "Ineligible Students Report", "ineligible-students",
                                     date_from, date_to, format)


async def _fetch_incident_reports(db: AsyncSession, date_from, date_to) -> list[dict]:
    reporter = aliased(User)
    student = aliased(User)
    query = (
        select(Incident, reporter, student)
        .join(reporter, reporter.id == Incident.reported_by_id)
        .outerjoin(student, student.id == Incident.involved_student_id)
    )
    if date_from:
        query = query.where(Incident.incident_date >= date_from)
    if date_to:
        query = query.where(Incident.incident_date <= date_to)
    query = query.order_by(Incident.incident_date.desc())
    result = await db.execute(query)
    rows = []
    for inc, reporter_u, student_u in result.all():
        rows.append({
            "Title": inc.title,
            "Category": inc.category.value,
            "Severity": inc.severity.value,
            "Status": inc.status.value,
            "Incident Date": inc.incident_date,
            "Location": inc.location,
            "Reported By": reporter_u.full_name,
            "Involved Student": student_u.full_name if student_u else "",
            "Resolution": inc.resolution,
        })
    return rows


async def _fetch_incident_categories(db: AsyncSession, date_from, date_to) -> list[dict]:
    query = select(Incident)
    if date_from:
        query = query.where(Incident.incident_date >= date_from)
    if date_to:
        query = query.where(Incident.incident_date <= date_to)
    result = await db.execute(query)
    stats: dict[str, dict] = {}
    for inc in result.scalars().all():
        entry = stats.setdefault(inc.category.value, {
            "Category": inc.category.value.title(),
            "Total": 0, "Open": 0, "Under Review": 0, "Resolved": 0, "Closed": 0,
        })
        entry["Total"] += 1
        key = inc.status.value
        if key == "open":
            entry["Open"] += 1
        elif key == "under_review":
            entry["Under Review"] += 1
        elif key == "resolved":
            entry["Resolved"] += 1
        elif key == "closed":
            entry["Closed"] += 1
    return sorted(stats.values(), key=lambda x: x["Total"], reverse=True)


async def _fetch_incident_summary(db: AsyncSession, date_from, date_to) -> list[dict]:
    query = select(Incident)
    if date_from:
        query = query.where(Incident.incident_date >= date_from)
    if date_to:
        query = query.where(Incident.incident_date <= date_to)
    result = await db.execute(query)
    incidents = result.scalars().all()
    status_counts: dict[str, int] = {"open": 0, "under_review": 0, "resolved": 0, "closed": 0}
    severity_counts: dict[str, int] = {"low": 0, "medium": 0, "high": 0, "critical": 0}
    for inc in incidents:
        status_counts[inc.status.value] = status_counts.get(inc.status.value, 0) + 1
        severity_counts[inc.severity.value] = severity_counts.get(inc.severity.value, 0) + 1
    return [
        {"Metric": "Total Incidents", "Value": len(incidents)},
        {"Metric": "Open", "Value": status_counts["open"]},
        {"Metric": "Under Review", "Value": status_counts["under_review"]},
        {"Metric": "Resolved", "Value": status_counts["resolved"]},
        {"Metric": "Closed", "Value": status_counts["closed"]},
        {"Metric": "Low Severity", "Value": severity_counts["low"]},
        {"Metric": "Medium Severity", "Value": severity_counts["medium"]},
        {"Metric": "High Severity", "Value": severity_counts["high"]},
        {"Metric": "Critical Severity", "Value": severity_counts["critical"]},
    ]


async def _incident_report(
    _user: NotStudent,
    db: AsyncSession,
    fetch,
    title: str,
    slug: str,
    date_from: date | None,
    date_to: date | None,
    format: str,
) -> StreamingResponse | list:
    rows = await fetch(db, _day_start(date_from), _day_end(date_to))
    headers = ["Title", "Category", "Severity", "Status", "Incident Date", "Location",
               "Reported By", "Involved Student", "Resolution"]
    return await _export_rows(db, _user, rows, headers, title,
                              _range_label(date_from, date_to), format, slug, slug, date_from, date_to)


@router.get("/incidents/reports", response_model=None, summary="Incident reports log")
async def incident_reports_log(
    _user: NotStudent,
    db: Annotated[AsyncSession, Depends(get_db)],
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    format: str = Query("json", pattern="^(json|csv|xlsx|pdf)$"),
) -> StreamingResponse | list:
    return await _incident_report(_user, db, _fetch_incident_reports,
                                  "Incident Reports", "incident-reports", date_from, date_to, format)


@router.get("/incidents/categories", response_model=None, summary="Incident categories report")
async def incident_categories_log(
    _user: NotStudent,
    db: Annotated[AsyncSession, Depends(get_db)],
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    format: str = Query("json", pattern="^(json|csv|xlsx|pdf)$"),
) -> StreamingResponse | list:
    rows = await _fetch_incident_categories(db, _day_start(date_from), _day_end(date_to))
    headers = ["Category", "Total", "Open", "Under Review", "Resolved", "Closed"]
    return await _export_rows(db, _user, rows, headers, "Incident Categories Report",
                              _range_label(date_from, date_to), format,
                              "incident-categories", "incident-categories", date_from, date_to)


@router.get("/incidents/summary", response_model=None, summary="Incident summary report")
async def incident_summary_log(
    _user: NotStudent,
    db: Annotated[AsyncSession, Depends(get_db)],
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    format: str = Query("json", pattern="^(json|csv|xlsx|pdf)$"),
) -> StreamingResponse | list:
    rows = await _fetch_incident_summary(db, _day_start(date_from), _day_end(date_to))
    headers = ["Metric", "Value"]
    return await _export_rows(db, _user, rows, headers, "Incident Summary Report",
                              _range_label(date_from, date_to), format,
                              "incident-summary", "incident-summary", date_from, date_to)


async def _fetch_sanctions(db: AsyncSession, statuses: list | None, date_from, date_to) -> list[dict]:
    student = aliased(User)
    issuer = aliased(User)
    query = (
        select(Sanction, student, issuer)
        .join(student, student.id == Sanction.student_id)
        .join(issuer, issuer.id == Sanction.issued_by_id)
    )
    if statuses:
        query = query.where(Sanction.status.in_(statuses))
    if date_from:
        query = query.where(Sanction.created_at >= date_from)
    if date_to:
        query = query.where(Sanction.created_at <= date_to)
    query = query.order_by(Sanction.created_at.desc())
    result = await db.execute(query)
    rows = []
    for s, stud, iss in result.all():
        rows.append({
            "Student Name": stud.full_name,
            "Student ID": stud.student_id or "",
            "Violation": s.violation_type.value,
            "Severity": s.severity.value,
            "Status": s.status.value,
            "Violation Date": s.violation_date,
            "Start Date": s.start_date,
            "End Date": s.end_date,
            "Issued By": iss.full_name,
            "Penalty": s.penalty,
            "Compliant": s.is_compliant,
            "Acknowledged": s.acknowledged_by_student,
        })
    return rows


async def _sanction_report(
    _user: NotStudent,
    db: AsyncSession,
    statuses: list | None,
    title: str,
    slug: str,
    date_from: date | None,
    date_to: date | None,
    format: str,
) -> StreamingResponse | list:
    rows = await _fetch_sanctions(db, statuses, _day_start(date_from), _day_end(date_to))
    headers = ["Student Name", "Student ID", "Violation", "Severity", "Status",
               "Violation Date", "Start Date", "End Date", "Issued By", "Penalty",
               "Compliant", "Acknowledged"]
    return await _export_rows(db, _user, rows, headers, title,
                              _range_label(date_from, date_to), format, slug, slug, date_from, date_to)


@router.get("/sanctions/active", response_model=None, summary="Active sanctions report")
async def active_sanctions_report(
    _user: NotStudent,
    db: Annotated[AsyncSession, Depends(get_db)],
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    format: str = Query("json", pattern="^(json|csv|xlsx|pdf)$"),
) -> StreamingResponse | list:
    return await _sanction_report(_user, db, [SanctionStatus.ACTIVE],
                                  "Active Sanctions Report", "active-sanctions",
                                  date_from, date_to, format)


@router.get("/sanctions/completed", response_model=None, summary="Completed sanctions report")
async def completed_sanctions_report(
    _user: NotStudent,
    db: Annotated[AsyncSession, Depends(get_db)],
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    format: str = Query("json", pattern="^(json|csv|xlsx|pdf)$"),
) -> StreamingResponse | list:
    return await _sanction_report(_user, db, [SanctionStatus.SERVED],
                                  "Completed Sanctions Report", "completed-sanctions",
                                  date_from, date_to, format)


@router.get("/sanctions/history", response_model=None, summary="Sanction history report")
async def sanction_history_report(
    _user: NotStudent,
    db: Annotated[AsyncSession, Depends(get_db)],
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    format: str = Query("json", pattern="^(json|csv|xlsx|pdf)$"),
) -> StreamingResponse | list:
    return await _sanction_report(_user, db, None,
                                  "Sanction History Report", "sanction-history",
                                  date_from, date_to, format)
