import csv
import io
import re
import uuid
from datetime import UTC, datetime
from typing import Annotated

import openpyxl
import structlog
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.dependencies import CurrentUser, get_db
from app.core.exceptions import NotFoundError
from app.models.audit import AuditLog
from app.models.user import UserRole
from app.schemas.audit import AuditLogRead
from app.schemas.common import PaginatedResponse
from app.services.audit_service import audit_log

logger = structlog.get_logger(__name__)
router = APIRouter()

_EDITOR_ROLES = {UserRole.ADMIN}


def _require_editor(current_user: CurrentUser) -> None:
    if current_user.role not in _EDITOR_ROLES:
        from app.core.exceptions import ForbiddenError
        raise ForbiddenError("Only admin can access audit logs.")


def _build_filters(
    search: str | None,
    module: str | None,
    action: str | None,
    status: str | None,
    user_id: uuid.UUID | None,
    ip_address: str | None,
    date_from: datetime | None,
    date_to: datetime | None,
) -> list:
    filters = []
    if search:
        like = f"%{search}%"
        filters.append(
            AuditLog.action.ilike(like)
            | AuditLog.description.ilike(like)
            | AuditLog.admin_name.ilike(like)
            | AuditLog.admin_email.ilike(like)
            | AuditLog.resource_type.ilike(like)
            | AuditLog.resource_id.ilike(like)
        )
    if module:
        filters.append(AuditLog.module == module)
    if action:
        filters.append(AuditLog.action == action)
    if status:
        filters.append(AuditLog.status == status)
    if user_id:
        filters.append(AuditLog.user_id == user_id)
    if ip_address:
        filters.append(AuditLog.ip_address == ip_address)
    if date_from:
        filters.append(AuditLog.created_at >= date_from)
    if date_to:
        filters.append(AuditLog.created_at <= date_to)
    return filters


def _log_to_read(log: AuditLog) -> AuditLogRead:
    return AuditLogRead.model_validate(log)


@router.get(
    "",
    response_model=PaginatedResponse[AuditLogRead],
    summary="List audit logs (Admin / Director)",
)
async def list_audit_logs(
    _editor: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    search: str | None = Query(None),
    module: str | None = Query(None),
    action: str | None = Query(None),
    status: str | None = Query(None),
    user_id: uuid.UUID | None = Query(None),
    ip_address: str | None = Query(None),
    date_from: datetime | None = Query(None),
    date_to: datetime | None = Query(None),
    sort_order: str = Query("desc", regex="^(asc|desc)$"),
) -> PaginatedResponse[AuditLogRead]:
    _require_editor(_editor)

    filters = _build_filters(search, module, action, status, user_id, ip_address, date_from, date_to)

    count_q = select(func.count()).select_from(AuditLog)
    if filters:
        count_q = count_q.where(and_(*filters))
    total = (await db.execute(count_q)).scalar() or 0

    order = AuditLog.created_at.desc() if sort_order == "desc" else AuditLog.created_at.asc()
    q = (
        select(AuditLog)
        .where(and_(*filters)) if filters else select(AuditLog)
    )
    q = q.order_by(order).offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(q)
    logs = result.scalars().all()

    pages = max(1, -(-total // page_size))
    return PaginatedResponse(
        items=[_log_to_read(l) for l in logs],
        total=total,
        page=page,
        page_size=page_size,
        pages=pages,
    )


@router.get(
    "/{log_id}",
    response_model=AuditLogRead,
    summary="Get audit log detail (Admin / Director)",
)
async def get_audit_log(
    log_id: uuid.UUID,
    _editor: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> AuditLogRead:
    _require_editor(_editor)

    log = await db.get(AuditLog, log_id)
    if not log:
        raise NotFoundError("Audit log", str(log_id))

    await audit_log(
        db=db,
        action="AUDIT_LOG_VIEWED",
        module="Audit Logs",
        description=f"Viewed audit log detail {log_id}",
        resource_type="AuditLog",
        resource_id=str(log_id),
        current_user=_editor,
    )

    return _log_to_read(log)


@router.get(
    "/filters/modules",
    summary="Get distinct module values for filter dropdown",
)
async def get_distinct_modules(
    _editor: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> list[str]:
    _require_editor(_editor)
    result = await db.execute(
        select(AuditLog.module).where(AuditLog.module.isnot(None)).distinct()
    )
    return sorted([r[0] for r in result.all() if r[0]])


@router.get(
    "/filters/actions",
    summary="Get distinct action values for filter dropdown",
)
async def get_distinct_actions(
    _editor: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> list[str]:
    _require_editor(_editor)
    result = await db.execute(
        select(AuditLog.action).distinct()
    )
    return sorted([r[0] for r in result.all() if r[0]])


@router.get(
    "/export/csv",
    summary="Export audit logs to CSV",
)
async def export_audit_logs_csv(
    _editor: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    search: str | None = Query(None),
    module: str | None = Query(None),
    action: str | None = Query(None),
    status: str | None = Query(None),
    user_id: uuid.UUID | None = Query(None),
    ip_address: str | None = Query(None),
    date_from: datetime | None = Query(None),
    date_to: datetime | None = Query(None),
) -> StreamingResponse:
    _require_editor(_editor)

    filters = _build_filters(search, module, action, status, user_id, ip_address, date_from, date_to)
    q = select(AuditLog).order_by(AuditLog.created_at.desc())
    if filters:
        q = q.where(and_(*filters))
    q = q.limit(10000)
    result = await db.execute(q)
    logs = result.scalars().all()

    await audit_log(
        db=db, action="AUDIT_LOG_EXPORTED", module="Audit Logs",
        description=f"Exported {len(logs)} audit logs to CSV",
        current_user=_editor,
    )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "ID", "Timestamp", "Admin Name", "Email", "Role", "Action", "Module",
        "Description", "Resource Type", "Resource ID", "Status", "Failure Reason",
        "IP Address", "Browser", "OS", "Device", "Request URL", "HTTP Method",
    ])
    for log in logs:
        writer.writerow([
            str(log.id), log.created_at.isoformat() if log.created_at else "",
            log.admin_name or "", log.admin_email or "", log.admin_role or "",
            log.action, log.module or "", log.description or "",
            log.resource_type or "", log.resource_id or "", log.status,
            log.failure_reason or "", log.ip_address or "",
            log.browser or "", log.os or "", log.device_info or "",
            log.request_url or "", log.http_method or "",
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=audit-logs-{datetime.now(UTC).strftime('%Y%m%d')}.csv"},
    )


@router.get(
    "/export/xlsx",
    summary="Export audit logs to Excel",
)
async def export_audit_logs_xlsx(
    _editor: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    search: str | None = Query(None),
    module: str | None = Query(None),
    action: str | None = Query(None),
    status: str | None = Query(None),
    user_id: uuid.UUID | None = Query(None),
    ip_address: str | None = Query(None),
    date_from: datetime | None = Query(None),
    date_to: datetime | None = Query(None),
) -> StreamingResponse:
    _require_editor(_editor)

    filters = _build_filters(search, module, action, status, user_id, ip_address, date_from, date_to)
    q = select(AuditLog).order_by(AuditLog.created_at.desc())
    if filters:
        q = q.where(and_(*filters))
    q = q.limit(10000)
    result = await db.execute(q)
    logs = result.scalars().all()

    await audit_log(
        db=db, action="AUDIT_LOG_EXPORTED", module="Audit Logs",
        description=f"Exported {len(logs)} audit logs to Excel",
        current_user=_editor,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Logs"

    header_fill = openpyxl.styles.PatternFill("solid", fgColor="1E3A5F")
    header_font = openpyxl.styles.Font(color="FFFFFF", bold=True)

    headers = [
        "ID", "Timestamp", "Admin", "Email", "Role", "Action", "Module",
        "Description", "Resource Type", "Resource ID", "Status",
        "IP Address", "Browser", "OS", "Device", "URL", "Method",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row, log in enumerate(logs, 2):
        ws.cell(row=row, column=1, value=str(log.id))
        ws.cell(row=row, column=2, value=log.created_at.isoformat() if log.created_at else "")
        ws.cell(row=row, column=3, value=log.admin_name or "")
        ws.cell(row=row, column=4, value=log.admin_email or "")
        ws.cell(row=row, column=5, value=log.admin_role or "")
        ws.cell(row=row, column=6, value=log.action)
        ws.cell(row=row, column=7, value=log.module or "")
        ws.cell(row=row, column=8, value=log.description or "")
        ws.cell(row=row, column=9, value=log.resource_type or "")
        ws.cell(row=row, column=10, value=log.resource_id or "")
        ws.cell(row=row, column=11, value=log.status)
        ws.cell(row=row, column=12, value=log.ip_address or "")
        ws.cell(row=row, column=13, value=log.browser or "")
        ws.cell(row=row, column=14, value=log.os or "")
        ws.cell(row=row, column=15, value=log.device_info or "")
        ws.cell(row=row, column=16, value=log.request_url or "")
        ws.cell(row=row, column=17, value=log.http_method or "")

    for col in range(1, len(headers) + 1):
        max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(len(logs) + 2, 50)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=audit-logs-{datetime.now(UTC).strftime('%Y%m%d')}.xlsx"},
    )


_ISO_DT_RE = re.compile(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}(?:\.\d+)?(?:Z|[+-]\d{2}:\d{2})?")


def _fmt_embedded_datetimes(text: str) -> str:
    def _repl(m: re.Match) -> str:
        try:
            dt = datetime.fromisoformat(m.group(0).replace("Z", "+00:00"))
        except ValueError:
            return m.group(0)
        if dt.tzinfo is not None:
            dt = dt.astimezone(UTC)
        return dt.strftime("%B %d, %Y %I:%M %p")
    return _ISO_DT_RE.sub(_repl, text)


@router.get(
    "/export/pdf",
    summary="Export audit logs to PDF",
)
async def export_audit_logs_pdf(
    _editor: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    search: str | None = Query(None),
    module: str | None = Query(None),
    action: str | None = Query(None),
    status: str | None = Query(None),
    user_id: uuid.UUID | None = Query(None),
    ip_address: str | None = Query(None),
    date_from: datetime | None = Query(None),
    date_to: datetime | None = Query(None),
) -> StreamingResponse:
    _require_editor(_editor)

    from xhtml2pdf import pisa

    filters = _build_filters(search, module, action, status, user_id, ip_address, date_from, date_to)
    q = select(AuditLog).order_by(AuditLog.created_at.desc())
    if filters:
        q = q.where(and_(*filters))
    q = q.limit(5000)
    result = await db.execute(q)
    logs = result.scalars().all()

    await audit_log(
        db=db, action="AUDIT_LOG_EXPORTED", module="Audit Logs",
        description=f"Exported {len(logs)} audit logs to PDF",
        current_user=_editor,
    )

    rows_html = ""
    for log in logs:
        rows_html += f"""
        <tr>
            <td>{log.created_at.strftime('%B %d, %Y %I:%M %p') if log.created_at else ''}</td>
            <td>{log.admin_name or ''}</td>
            <td>{log.action}</td>
            <td>{log.module or ''}</td>
            <td>{log.status}</td>
            <td>{log.ip_address or ''}</td>
            <td>{_fmt_embedded_datetimes(log.description or '')}</td>
        </tr>"""

    html = f"""
    <html><head><style>
        @page {{ size: A4 landscape; margin: 12mm 9mm; }}
        body {{ font-family: Helvetica, sans-serif; font-size: 9pt; }}
        h1 {{ font-size: 16pt; color: #1E3A5F; }}
        p {{ color: #666; font-size: 8pt; }}
        table {{ width: 100%; table-layout: fixed; border-collapse: collapse; margin-top: 10px; }}
        th {{ background: #1E3A5F; color: white; padding: 5px 6px; text-align: left; font-size: 8pt; word-wrap: break-word; }}
        td {{ padding: 4px 6px; border-bottom: 1px solid #ddd; font-size: 8pt; word-wrap: break-word; }}
        tr:nth-child(even) {{ background: #f5f7fa; }}
    </style></head><body>
        <h1>OSCA Audit Log Report</h1>
        <p>Generated: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')} | Records: {len(logs)}</p>
        <table>
            <tr><th>Date</th><th>Admin</th><th>Action</th><th>Module</th><th>Status</th><th>IP</th><th>Description</th></tr>
            {rows_html}
        </table>
    </body></html>"""

    output = io.BytesIO()
    pisa.CreatePDF(io.BytesIO(html.encode("utf-8")), dest=output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=audit-logs-{datetime.now(UTC).strftime('%Y%m%d')}.pdf"},
    )
